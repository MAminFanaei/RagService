"""
XLSX parser using openpyxl.

Extracts:
  - Each worksheet as a section (sheet name = section title)
  - Tables as Markdown elements (full sheet data)
  - Named tables within sheets as separate elements
  - Page number = sheet index (1-based)
"""

from __future__ import annotations

import io
from typing import Any

import structlog

from ingestion.parsers.base import BaseParser, ParsedElement, ParserError

logger = structlog.get_logger(__name__)


def _rows_to_markdown(rows: list[list[str]]) -> tuple[str, str]:
    """Return (plain_text, markdown) from a 2D list of cell strings."""
    if not rows:
        return "", ""

    plain = "\n".join(" | ".join(r) for r in rows)

    try:
        from tabulate import tabulate  # type: ignore

        if len(rows) > 1:
            md = tabulate(rows[1:], headers=rows[0], tablefmt="pipe")
        else:
            md = tabulate(rows, tablefmt="pipe")
    except ImportError:
        if len(rows) > 1:
            header = " | ".join(rows[0])
            sep = " | ".join(["---"] * len(rows[0]))
            body = "\n".join(" | ".join(r) for r in rows[1:])
            md = f"{header}\n{sep}\n{body}"
        else:
            md = plain

    return plain, md


def _cell_str(cell) -> str:
    """Convert an openpyxl cell value to string, stripping None."""
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


class XlsxParser(BaseParser):
    """Parser for Excel .xlsx files."""

    @property
    def name(self) -> str:
        return "xlsx"

    async def parse(self, data: bytes, filename: str) -> list[ParsedElement]:
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:
            raise ParserError(self.name, "openpyxl not installed", exc)

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(data), read_only=True, data_only=True
            )
        except Exception as exc:
            raise ParserError(self.name, f"Cannot open XLSX: {exc}", exc)

        elements: list[ParsedElement] = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            page_num = sheet_idx + 1
            ws = wb[sheet_name]

            # ---- Sheet heading ----
            elements.append(
                ParsedElement(
                    text=sheet_name,
                    parser_name=self.name,
                    element_type="heading",
                    heading_level=1,
                    page_number=page_num,
                    section_title=sheet_name,
                    section_path=[sheet_name],
                    raw_metadata={"sheet": sheet_name},
                )
            )

            # ---- Collect all non-empty rows ----
            all_rows: list[list[str]] = []
            for row in ws.iter_rows():
                row_data = [_cell_str(c) for c in row]
                # Skip completely empty rows
                if any(v for v in row_data):
                    all_rows.append(row_data)

            if not all_rows:
                continue

            # ---- Remove trailing empty columns across all rows ----
            if all_rows:
                max_cols = max(len(r) for r in all_rows)
                # Find last column index that has any data
                last_col = 0
                for r in all_rows:
                    for ci in range(len(r) - 1, -1, -1):
                        if r[ci]:
                            last_col = max(last_col, ci)
                all_rows = [r[: last_col + 1] for r in all_rows]

            plain, md = _rows_to_markdown(all_rows)

            if plain.strip():
                elements.append(
                    ParsedElement(
                        text=plain,
                        parser_name=self.name,
                        element_type="table",
                        is_table=True,
                        table_markdown=md,
                        page_number=page_num,
                        section_title=sheet_name,
                        section_path=[sheet_name],
                        raw_metadata={
                            "sheet": sheet_name,
                            "row_count": len(all_rows),
                        },
                    )
                )

        wb.close()

        logger.info(
            "xlsx_parsed",
            filename=filename,
            sheet_count=len(wb.sheetnames),
            element_count=len(elements),
        )
        return elements
