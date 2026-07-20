# test_quality_xlsx.py

import json
import time
from pathlib import Path

def save_result(output_path, method, data):
    with open(output_path / f"{method}.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  ✅ {method:<20} → {output_path / f'{method}.json'}")

def run_xlsx_tests(file_path: Path, output_dir: Path, args):
    print(f"📊 Testing XLSX parser on: {file_path.name}\n")
    
    print("▶ Testing openpyxl...")
    try:
        from openpyxl import load_workbook
        from tabulate import tabulate
        
        t0 = time.perf_counter()
        wb = load_workbook(file_path, data_only=True)
        
        sheets_data = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Get all rows
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Skip empty rows
                if any(cell is not None for cell in row):
                    rows.append([str(cell) if cell is not None else "" for cell in row])
            
            if not rows:
                continue
            
            # Convert to markdown table
            markdown = tabulate(rows[1:], headers=rows[0], tablefmt="github") if len(rows) > 1 else str(rows)
            
            sheets_data.append({
                "name": sheet_name,
                "row_count": len(rows),
                "col_count": len(rows[0]) if rows else 0,
                "markdown_preview": markdown[:2000],
                "raw_rows_preview": rows[:10],
            })
        
        elapsed = time.perf_counter() - t0
        
        result = {
            "parser": "openpyxl",
            "file": file_path.name,
            "time_seconds": round(elapsed, 3),
            "sheet_count": len(sheets_data),
            "sheets": sheets_data,
            "verdict": {
                "quality": f"✅ {len(sheets_data)} sheets extracted",
            }
        }
        save_result(output_dir, "parser_openpyxl", result)
        
    except ImportError:
        print(f"  ⏭️  openpyxl not installed — skip")
    except Exception as e:
        print(f"  ❌ openpyxl failed: {e}")